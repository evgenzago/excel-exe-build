import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import time
import os
import threading


class UltraModernApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Report Generator")
        self.root.geometry("1100x750")
        self.root.configure(bg='#0f0f0f')
        
        # Modern dark theme colors
        self.theme = {
            'bg_primary': '#0f0f0f',
            'bg_secondary': '#1a1a1a',
            'bg_card': '#242424',
            'bg_hover': '#2d2d2d',
            'accent': '#6C63FF',
            'accent_light': '#8B83FF',
            'accent_dark': '#5A52D5',
            'success': '#00D9A5',
            'warning': '#FFB800',
            'error': '#FF6B6B',
            'text_primary': '#FFFFFF',
            'text_secondary': '#A0A0A0',
            'text_muted': '#666666',
            'border': '#333333',
            'gradient_start': '#6C63FF',
            'gradient_end': '#4ECDC4'
        }
        
        self.file_path = None
        self.start_time = None
        self.running = False
        
        # Custom fonts
        self.font_title = ('SF Pro Display', 28, 'bold')
        self.font_subtitle = ('SF Pro Display', 13)
        self.font_heading = ('SF Pro Display', 15, 'bold')
        self.font_body = ('SF Pro Text', 11)
        self.font_mono = ('SF Mono', 10)
        self.font_button = ('SF Pro Text', 12, 'bold')
        
        self.setup_window()
        self.create_ui()
    
    def setup_window(self):
        """Configure window appearance"""
        self.root.resizable(True, True)
        
        # Try to set app icon (optional)
        try:
            self.root.iconbitmap(default='')
        except:
            pass
    
    def create_ui(self):
        """Create the main UI layout"""
        # Main container
        main_container = tk.Frame(self.root, bg=self.theme['bg_primary'])
        main_container.pack(fill=tk.BOTH, expand=True, padx=30, pady=25)
        
        # Header section
        self.create_header(main_container)
        
        # Content area
        content_area = tk.Frame(main_container, bg=self.theme['bg_primary'])
        content_area.pack(fill=tk.BOTH, expand=True, pady=20)
        
        # Left panel - Controls
        left_panel = tk.Frame(content_area, bg=self.theme['bg_primary'], width=350)
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 15))
        left_panel.pack_propagate(False)
        
        # Right panel - Logs
        right_panel = tk.Frame(content_area, bg=self.theme['bg_primary'])
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        # Create sections in left panel
        self.create_file_section(left_panel)
        self.create_progress_section(left_panel)
        self.create_action_section(left_panel)
        
        # Create logs section in right panel
        self.create_logs_section(right_panel)
    
    def create_header(self, parent):
        """Create modern header with gradient effect simulation"""
        header_frame = tk.Frame(parent, bg=self.theme['bg_primary'])
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Title with accent color
        title_label = tk.Label(
            header_frame,
            text="Excel Report Generator",
            font=self.font_title,
            fg=self.theme['accent'],
            bg=self.theme['bg_primary']
        )
        title_label.pack(anchor='w')
        
        # Subtitle
        subtitle_label = tk.Label(
            header_frame,
            text="Automated data processing and report generation",
            font=self.font_subtitle,
            fg=self.theme['text_secondary'],
            bg=self.theme['bg_primary']
        )
        subtitle_label.pack(anchor='w', pady=(5, 0))
        
        # Separator line
        separator = tk.Frame(header_frame, height=2, bg=self.theme['border'])
        separator.pack(fill=tk.X, pady=(15, 0))
    
    def create_file_section(self, parent):
        """Create file selection card"""
        card = tk.Frame(parent, bg=self.theme['bg_card'], highlightthickness=0)
        card.pack(fill=tk.X, pady=(0, 15))
        
        # Card padding
        card_inner = tk.Frame(card, bg=self.theme['bg_card'], padx=20, pady=18)
        card_inner.pack(fill=tk.X, expand=True)
        
        # Section title
        tk.Label(
            card_inner,
            text="FILE SELECTION",
            font=('SF Pro Text', 10, 'bold'),
            fg=self.theme['text_muted'],
            bg=self.theme['bg_card']
        ).pack(anchor='w', pady=(0, 12))
        
        # File display area
        file_display = tk.Frame(card_inner, bg=self.theme['bg_secondary'], 
                               highlightbackground=self.theme['border'],
                               highlightthickness=1)
        file_display.pack(fill=tk.X, pady=(0, 12))
        
        self.file_var = tk.StringVar(value="No file selected")
        file_label = tk.Label(
            file_display,
            textvariable=self.file_var,
            font=self.font_body,
            fg=self.theme['text_secondary'],
            bg=self.theme['bg_secondary'],
            anchor='w',
            padx=12,
            pady=10,
            wraplength=280
        )
        file_label.pack(fill=tk.X)
        
        # Browse button
        browse_btn = tk.Button(
            card_inner,
            text="Browse Files",
            command=self.browse_file,
            font=self.font_button,
            bg=self.theme['accent'],
            fg=self.theme['text_primary'],
            activebackground=self.theme['accent_light'],
            activeforeground=self.theme['text_primary'],
            relief=tk.FLAT,
            padx=20,
            pady=10,
            cursor='hand2',
            borderwidth=0
        )
        browse_btn.pack(fill=tk.X)
        
        # Bind hover effects
        browse_btn.bind('<Enter>', lambda e: browse_btn.config(bg=self.theme['accent_light']))
        browse_btn.bind('<Leave>', lambda e: browse_btn.config(bg=self.theme['accent']))
    
    def create_progress_section(self, parent):
        """Create progress indicator card"""
        card = tk.Frame(parent, bg=self.theme['bg_card'], highlightthickness=0)
        card.pack(fill=tk.X, pady=(0, 15))
        
        card_inner = tk.Frame(card, bg=self.theme['bg_card'], padx=20, pady=18)
        card_inner.pack(fill=tk.X, expand=True)
        
        # Header with timer
        header_frame = tk.Frame(card_inner, bg=self.theme['bg_card'])
        header_frame.pack(fill=tk.X, pady=(0, 12))
        
        tk.Label(
            header_frame,
            text="PROGRESS",
            font=('SF Pro Text', 10, 'bold'),
            fg=self.theme['text_muted'],
            bg=self.theme['bg_card']
        ).pack(side=tk.LEFT)
        
        self.timer_var = tk.StringVar(value="00:00:00")
        tk.Label(
            header_frame,
            textvariable=self.timer_var,
            font=('SF Mono', 11, 'bold'),
            fg=self.theme['accent'],
            bg=self.theme['bg_card']
        ).pack(side=tk.RIGHT)
        
        # Progress bar container
        progress_bg = tk.Frame(card_inner, bg=self.theme['bg_secondary'],
                              height=8, highlightthickness=0)
        progress_bg.pack(fill=tk.X, pady=(0, 8))
        
        # Actual progress bar
        self.progress_bar = tk.Canvas(progress_bg, height=8, bg=self.theme['bg_secondary'],
                                     highlightthickness=0, bd=0)
        self.progress_bar.pack(fill=tk.X, expand=True)
        
        # Fill rectangle (will be updated)
        self.progress_rect = self.progress_bar.create_rectangle(
            0, 0, 0, 8, fill=self.theme['accent'], outline=''
        )
        
        # Percentage label
        self.progress_var = tk.StringVar(value="0%")
        tk.Label(
            card_inner,
            textvariable=self.progress_var,
            font=('SF Pro Text', 10),
            fg=self.theme['text_secondary'],
            bg=self.theme['bg_card']
        ).pack(anchor='e')
    
    def create_action_section(self, parent):
        """Create action buttons card"""
        card = tk.Frame(parent, bg=self.theme['bg_card'], highlightthickness=0)
        card.pack(fill=tk.X, pady=(0, 15))
        
        card_inner = tk.Frame(card, bg=self.theme['bg_card'], padx=20, pady=18)
        card_inner.pack(fill=tk.X, expand=True)
        
        tk.Label(
            card_inner,
            text="ACTIONS",
            font=('SF Pro Text', 10, 'bold'),
            fg=self.theme['text_muted'],
            bg=self.theme['bg_card']
        ).pack(anchor='w', pady=(0, 12))
        
        # Start button
        self.start_btn = tk.Button(
            card_inner,
            text="▶  Start Processing",
            command=self.start_processing,
            font=self.font_button,
            bg=self.theme['success'],
            fg=self.theme['text_primary'],
            activebackground='#00E6B8',
            activeforeground=self.theme['text_primary'],
            relief=tk.FLAT,
            padx=20,
            pady=12,
            cursor='hand2',
            borderwidth=0
        )
        self.start_btn.pack(fill=tk.X)
        
        self.start_btn.bind('<Enter>', lambda e: self.start_btn.config(bg='#00E6B8'))
        self.start_btn.bind('<Leave>', lambda e: self.start_btn.config(bg=self.theme['success']))
        
        # Status indicator
        status_frame = tk.Frame(card_inner, bg=self.theme['bg_card'])
        status_frame.pack(fill=tk.X, pady=(12, 0))
        
        self.status_indicator = tk.Canvas(status_frame, width=10, height=10,
                                         bg=self.theme['bg_card'], highlightthickness=0)
        self.status_indicator.pack(side=tk.LEFT)
        
        # Draw initial status circle
        self.status_circle = self.status_indicator.create_oval(
            2, 2, 8, 8, fill=self.theme['success'], outline=''
        )
        
        self.status_var = tk.StringVar(value="Ready")
        tk.Label(
            status_frame,
            textvariable=self.status_var,
            font=self.font_body,
            fg=self.theme['success'],
            bg=self.theme['bg_card']
        ).pack(side=tk.LEFT, padx=(8, 0))
    
    def create_logs_section(self, parent):
        """Create logs panel"""
        card = tk.Frame(parent, bg=self.theme['bg_card'], highlightthickness=0)
        card.pack(fill=tk.BOTH, expand=True)
        
        card_inner = tk.Frame(card, bg=self.theme['bg_card'], padx=20, pady=18)
        card_inner.pack(fill=tk.BOTH, expand=True)
        
        # Header
        header_frame = tk.Frame(card_inner, bg=self.theme['bg_card'])
        header_frame.pack(fill=tk.X, pady=(0, 12))
        
        tk.Label(
            header_frame,
            text="EXECUTION LOGS",
            font=('SF Pro Text', 10, 'bold'),
            fg=self.theme['text_muted'],
            bg=self.theme['bg_card']
        ).pack(side=tk.LEFT)
        
        clear_btn = tk.Button(
            header_frame,
            text="Clear",
            command=self.clear_logs,
            font=('SF Pro Text', 9),
            bg=self.theme['bg_card'],
            fg=self.theme['text_secondary'],
            activeforeground=self.theme['accent'],
            relief=tk.FLAT,
            cursor='hand2',
            borderwidth=0,
            padx=8,
            pady=4
        )
        clear_btn.pack(side=tk.RIGHT)
        
        # Log text area
        log_container = tk.Frame(card_inner, bg=self.theme['bg_secondary'],
                                highlightbackground=self.theme['border'],
                                highlightthickness=1)
        log_container.pack(fill=tk.BOTH, expand=True)
        
        self.log_text = tk.Text(
            log_container,
            font=self.font_mono,
            bg=self.theme['bg_secondary'],
            fg=self.theme['text_primary'],
            insertbackground=self.theme['accent'],
            selectbackground=self.theme['accent'],
            selectforeground=self.theme['text_primary'],
            relief=tk.FLAT,
            highlightthickness=0,
            padx=15,
            pady=12,
            wrap=tk.WORD,
            spacing1=3,
            spacing2=3,
            spacing3=3
        )
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Custom scrollbar
        scrollbar = tk.Scrollbar(log_container, command=self.log_text.yview,
                                bg=self.theme['bg_secondary'],
                                troughcolor=self.theme['bg_secondary'],
                                activebackground=self.theme['accent'],
                                elementborderwidth=0,
                                width=12)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=scrollbar.set)
        
        # Configure scrollbar styling
        scrollbar.config(command=self.log_text.yview)
    
    def browse_file(self):
        """Open file dialog"""
        self.file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")],
            title="Select Excel File"
        )
        if self.file_path:
            filename = os.path.basename(self.file_path)
            self.file_var.set(filename)
            self.log(f"✓ File selected: {filename}", "success")
            self.update_status("File Loaded", self.theme['success'])
    
    def log(self, message, level="info"):
        """Add log message with color coding"""
        timestamp = time.strftime("%H:%M:%S")
        
        # Color based on level
        if level == "success":
            color = self.theme['success']
        elif level == "error":
            color = self.theme['error']
        elif level == "warning":
            color = self.theme['warning']
        else:
            color = self.theme['text_secondary']
        
        self.log_text.insert(tk.END, f"[{timestamp}] ", self.theme['text_muted'])
        self.log_text.insert(tk.END, f"{message}\n", color)
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def clear_logs(self):
        """Clear all logs"""
        self.log_text.delete(1.0, tk.END)
    
    def update_status(self, text, color):
        """Update status indicator"""
        self.status_var.set(text)
        self.status_indicator.itemconfig(self.status_circle, fill=color)
        for widget in self.status_indicator.master.winfo_children():
            if isinstance(widget, tk.Label):
                widget.config(fg=color)
    
    def update_timer(self):
        """Update elapsed time display"""
        if self.running:
            elapsed = time.time() - self.start_time
            hours, remainder = divmod(elapsed, 3600)
            minutes, seconds = divmod(remainder, 60)
            self.timer_var.set(f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}")
            self.root.after(1000, self.update_timer)
    
    def update_progress(self, value):
        """Update progress bar"""
        self.progress_var.set(f"{int(value)}%")
        
        # Update canvas rectangle
        width = self.progress_bar.winfo_width()
        if width > 1:
            fill_width = (value / 100.0) * width
            self.progress_bar.coords(self.progress_rect, 0, 0, fill_width, 8)
        
        self.root.update_idletasks()
    
    def start_processing(self):
        """Start the processing workflow"""
        if not self.file_path:
            messagebox.showerror("Error", "Please select an Excel file first")
            return
        
        self.running = True
        self.start_time = time.time()
        self.update_timer()
        self.clear_logs()
        
        self.log("=" * 50, "info")
        self.log("Starting file processing...", "info")
        self.update_status("Processing...", self.theme['warning'])
        self.start_btn.config(state=tk.DISABLED, bg=self.theme['text_muted'])
        
        # Run in background thread
        thread = threading.Thread(target=self.process_excel_file, daemon=True)
        thread.start()
    
    def process_excel_file(self):
        """Main processing logic"""
        try:
            self.log("Opening Excel file...", "info")
            wb = openpyxl.load_workbook(self.file_path)
            self.update_progress(5)
            
            # Validate sheets
            required_sheets = ['DB', 'OUT', 'IN', 'CARD']
            for sheet in required_sheets:
                if sheet not in wb.sheetnames:
                    raise Exception(f"Sheet '{sheet}' not found")
            
            self.log("Creating REPORT sheet...", "info")
            if 'REPORT' in wb.sheetnames:
                wb.remove(wb['REPORT'])
            
            report_sheet = wb.create_sheet('REPORT')
            
            # Styled headers
            headers = [
                "ИД", "Счет для поиска", "Дата открытия первого счета", "Кол-во счетов",
                "Количество карт", "Из них виртуальных", "ИСХ_СБП_сумма", "ИСХ_СБП_Кол-во",
                "Вх_СБП_сумма", "Вх_СБП_кол-во", "СБП кол-во операций", "SUM781", "CNT781",
                "SUM785", "CNT785", "р2р кол-во операций", "снятие налички_сумма",
                "снятие налички_колво", "взнос налички_сумма", "взнос налички_колво",
                "Вх_оборот", "Исх_оборот", "Количество переводов и операций с нал"
            ]
            
            header_fill = PatternFill(start_color=self.theme['accent'], 
                                     end_color=self.theme['accent'], 
                                     fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11, name='Arial')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col_num, header in enumerate(headers, 1):
                cell = report_sheet.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                report_sheet.column_dimensions[get_column_letter(col_num)].width = 22
            
            self.update_progress(10)
            
            # Process DB data
            self.log("Collecting data from DB...", "info")
            db_sheet = wb['DB']
            db_phones = set()
            db_accounts = set()
            
            for row in db_sheet.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[0]:
                    db_phones.add(str(row[0]))
                if row[1]:
                    db_accounts.add(str(row[1]))
            
            self.log(f"Found {len(db_phones)} phones, {len(db_accounts)} accounts", "success")
            
            # Collect phone numbers
            phone_numbers = set()
            out_sheet = wb['OUT']
            in_sheet = wb['IN']
            card_sheet = wb['CARD']
            
            for row in out_sheet.iter_rows(min_row=2, max_col=5, values_only=True):
                if row[4] and str(row[4]) in db_phones:
                    phone_numbers.add(str(row[4]))
            
            for row in in_sheet.iter_rows(min_row=2, max_col=5, values_only=True):
                if row[4] and str(row[4]) in db_phones:
                    phone_numbers.add(str(row[4]))
            
            for row in card_sheet.iter_rows(min_row=2, max_col=14, values_only=True):
                if row[13] and str(row[13]) in db_phones:
                    phone_numbers.add(str(row[13]))
            
            self.log(f"Unique phones for report: {len(phone_numbers)}", "success")
            self.update_progress(20)
            
            # Write phones to report
            self.log("Writing phone numbers...", "info")
            current_row = 2
            for phone in sorted(phone_numbers):
                report_sheet.cell(row=current_row, column=1, value=phone)
                current_row += 1
            
            self.update_progress(30)
            
            # Process account data
            self.log("Processing account data...", "info")
            phone_to_account = {}
            phone_to_date = {}
            phone_count = {}
            
            for row in db_sheet.iter_rows(min_row=2, values_only=True):
                phone = str(row[0]) if row[0] else None
                account = str(row[1]) if row[1] else None
                date = str(row[4]) if row[4] else None
                
                if phone:
                    if phone not in phone_to_account:
                        phone_to_account[phone] = account
                        phone_to_date[phone] = date
                    phone_count[phone] = phone_count.get(phone, 0) + 1
            
            for row in report_sheet.iter_rows(min_row=2, max_col=1):
                phone = str(row[0].value) if row[0].value else None
                if phone and phone in phone_to_account:
                    report_sheet.cell(row=row[0].row, column=2, value=phone_to_account[phone])
                    report_sheet.cell(row=row[0].row, column=3, value=phone_to_date[phone])
                    report_sheet.cell(row=row[0].row, column=4, value=phone_count.get(phone, 0))
            
            self.update_progress(40)
            
            # Process cards
            self.log("Processing card data...", "info")
            phone_card_count = {}
            phone_virtual_count = {}
            
            for row in card_sheet.iter_rows(min_row=2, values_only=True):
                phone = str(row[13]) if row[13] else None
                card_type = str(row[2]) if row[2] else None
                
                if phone and phone in phone_numbers:
                    phone_card_count[phone] = phone_card_count.get(phone, 0) + 1
                    if card_type and "virtual" in card_type.lower():
                        phone_virtual_count[phone] = phone_virtual_count.get(phone, 0) + 1
            
            for row in report_sheet.iter_rows(min_row=2, max_col=1):
                phone = str(row[0].value) if row[0].value else None
                if phone:
                    report_sheet.cell(row=row[0].row, column=5, value=phone_card_count.get(phone, 0))
                    report_sheet.cell(row=row[0].row, column=6, value=phone_virtual_count.get(phone, 0))
            
            self.update_progress(50)
            
            # Process SBP operations
            self.log("Processing SBP transactions...", "info")
            out_amounts, out_counts = {}, {}
            in_amounts, in_counts = {}, {}
            
            for row in out_sheet.iter_rows(min_row=2, values_only=True):
                phone = str(row[4]) if row[4] else None
                amount = float(row[5]) if row[5] else 0
                if phone and phone in phone_numbers:
                    out_amounts[phone] = out_amounts.get(phone, 0) + amount
                    out_counts[phone] = out_counts.get(phone, 0) + 1
            
            for row in in_sheet.iter_rows(min_row=2, values_only=True):
                phone = str(row[4]) if row[4] else None
                amount = float(row[5]) if row[5] else 0
                if phone and phone in phone_numbers:
                    in_amounts[phone] = in_amounts.get(phone, 0) + amount
                    in_counts[phone] = in_counts.get(phone, 0) + 1
            
            for row in report_sheet.iter_rows(min_row=2, max_col=1):
                phone = str(row[0].value) if row[0].value else None
                if phone:
                    report_sheet.cell(row=row[0].row, column=7, value=out_amounts.get(phone, 0))
                    report_sheet.cell(row=row[0].row, column=8, value=out_counts.get(phone, 0))
                    report_sheet.cell(row=row[0].row, column=9, value=in_amounts.get(phone, 0))
                    report_sheet.cell(row=row[0].row, column=10, value=in_counts.get(phone, 0))
                    total_sbp = out_counts.get(phone, 0) + in_counts.get(phone, 0)
                    report_sheet.cell(row=row[0].row, column=11, value=total_sbp)
            
            self.update_progress(70)
            
            # Process 781/785 operations
            self.log("Processing card operations (781/785)...", "info")
            sum781, cnt781, sum785, cnt785 = {}, {}, {}, {}
            
            for row in card_sheet.iter_rows(min_row=2, values_only=True):
                phone = str(row[13]) if row[13] else None
                if phone and phone in phone_numbers:
                    amt781 = float(row[3]) if row[3] else 0
                    amt785 = float(row[5]) if row[5] else 0
                    sum781[phone] = sum781.get(phone, 0) + amt781
                    cnt781[phone] = cnt781.get(phone, 0) + (1 if row[3] else 0)
                    sum785[phone] = sum785.get(phone, 0) + amt785
                    cnt785[phone] = cnt785.get(phone, 0) + (1 if row[5] else 0)
            
            for row in report_sheet.iter_rows(min_row=2, max_col=1):
                phone = str(row[0].value) if row[0].value else None
                if phone:
                    report_sheet.cell(row=row[0].row, column=12, value=sum781.get(phone, 0))
                    report_sheet.cell(row=row[0].row, column=13, value=cnt781.get(phone, 0))
                    report_sheet.cell(row=row[0].row, column=14, value=sum785.get(phone, 0))
                    report_sheet.cell(row=row[0].row, column=15, value=cnt785.get(phone, 0))
                    p2p = cnt781.get(phone, 0) + cnt785.get(phone, 0)
                    report_sheet.cell(row=row[0].row, column=16, value=p2p)
            
            self.update_progress(80)
            
            # Cash operations
            self.log("Processing cash operations...", "info")
            cash_withdrawal_sum, cash_withdrawal_cnt = {}, {}
            cash_deposit_sum, cash_deposit_cnt = {}, {}
            
            for row in card_sheet.iter_rows(min_row=2, values_only=True):
                phone = str(row[13]) if row[13] else None
                if phone and phone in phone_numbers:
                    withdrawal = float(row[9]) if row[9] else 0
                    deposit = float(row[11]) if row[11] else 0
                    cash_withdrawal_sum[phone] = cash_withdrawal_sum.get(phone, 0) + withdrawal
                    cash_deposit_sum[phone] = cash_deposit_sum.get(phone, 0) + deposit
                    if row[9]:
                        cash_withdrawal_cnt[phone] = cash_withdrawal_cnt.get(phone, 0) + 1
                    if row[11]:
                        cash_deposit_cnt[phone] = cash_deposit_cnt.get(phone, 0) + 1
            
            for row in report_sheet.iter_rows(min_row=2, max_col=1):
                phone = str(row[0].value) if row[0].value else None
                if phone:
                    report_sheet.cell(row=row[0].row, column=17, value=cash_withdrawal_sum.get(phone, 0))
                    report_sheet.cell(row=row[0].row, column=18, value=cash_withdrawal_cnt.get(phone, 0))
                    report_sheet.cell(row=row[0].row, column=19, value=cash_deposit_sum.get(phone, 0))
                    report_sheet.cell(row=row[0].row, column=20, value=cash_deposit_cnt.get(phone, 0))
            
            self.update_progress(90)
            
            # Calculate totals
            self.log("Calculating totals...", "info")
            for row in report_sheet.iter_rows(min_row=2):
                if row[0].value:
                    incoming = (row[8].value or 0) + (row[13].value or 0) + (row[18].value or 0)
                    outgoing = (row[6].value or 0) + (row[11].value or 0) + (row[16].value or 0)
                    total_ops = (row[10].value or 0) + (row[15].value or 0) + \
                               (row[17].value or 0) + (row[19].value or 0)
                    row[20].value = incoming
                    row[21].value = outgoing
                    row[22].value = total_ops
            
            # Set numeric types
            numeric_cols = [7, 9, 12, 14, 17, 19, 21, 22]
            for col in numeric_cols:
                for cell in report_sheet[get_column_letter(col)][1:]:
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value)
                        except ValueError:
                            pass
            
            self.update_progress(95)
            
            # Save
            self.log("Saving file...", "info")
            wb.save(self.file_path)
            self.log("✓ File saved successfully!", "success")
            
            elapsed = time.time() - self.start_time
            hours, remainder = divmod(elapsed, 3600)
            minutes, seconds = divmod(remainder, 60)
            time_str = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"
            
            self.log(f"✓ Processing completed in {time_str}", "success")
            self.update_status("Completed", self.theme['success'])
            messagebox.showinfo("Success", f"Report generated successfully!\nTime: {time_str}")
            
        except Exception as e:
            self.log(f"✗ Error: {str(e)}", "error")
            self.update_status("Error", self.theme['error'])
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
        finally:
            self.running = False
            self.start_btn.config(state=tk.NORMAL, bg=self.theme['success'])


def main():
    root = tk.Tk()
    app = UltraModernApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
